# -*- coding: utf-8 -*-
"""
Sertlik Takip Uygulamasi
PDF sertlik raporlarini surukle-birak ile Excel tablosuna otomatik ekler.
"""

import os
import re
import sys
import json
import subprocess
import datetime
from pathlib import Path

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
    DND_AVAILABLE = True
except ImportError:
    DND_AVAILABLE = False

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# pdfminer for text extraction (pure python, no external binary needed)
from pdfminer.high_level import extract_text as pdfminer_extract_text

LABELS = {"MÜSTERI:", "MUSTERI:", "PARTI NO:", "EBAT:", "TARIH:", "OPERATÖR:", "OPERATOR:"}

APP_NAME = "Sertlik Takip"
CONFIG_FILE = os.path.join(os.path.expanduser("~"), ".sertlik_takip_config.json")

HEADERS = [
    "Ebat",
    "Ortalama Sertlik (HV)",
    "Musteri / Parti No",
    "Operator",
    "Dosya Adi",
    "Dosya Degisiklik Tarihi",
    "Dosya Degisiklik Saati",
]

KNOWN_OPERATORS = {
    "VOLKAN ERTUGRUL", "UMUT ALPARSLAN YALÇIN", "UMUT ALPARSLAN YALCIN",
    "YUSUF BIRBEN", "MUHAMMED KAYMAK", "EREN YURDAGÜL", "EREN YURDAGUL",
    "ALAATTIN UYSAL", "SINAN DEMIRAL",
}


# ---------------------------------------------------------------------------
# Konfigurasyon (Excel dosya yolunu hatirla)
# ---------------------------------------------------------------------------

def load_config():
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def save_config(cfg):
    try:
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(cfg, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


# ---------------------------------------------------------------------------
# PDF Parsing
# ---------------------------------------------------------------------------




DATE_PATTERN = re.compile(r"^\d{2}\.\d{2}\.\d{4}$")
EBAT_PATTERN = re.compile(r"^\d+[.,]\d+\s*[xX*\-,]\s*\d+[.,]\d+")
HV_LABEL_PATTERN = re.compile(r"^HV\s*[\d,\.]+$")
NUMBER_PATTERN = re.compile(r"^\d+[,\.]\d+$")


def extract_pdf_text(pdf_path):
    try:
        return pdfminer_extract_text(pdf_path)
    except Exception:
        return ""


def _nonempty_lines(text):
    return [l.strip() for l in text.split("\n") if l.strip()]


def extract_hv_values(text):
    """HV satirlarinin hemen oncesindeki ardisik sayi blogunun son N elemanini
    sertlik degeri olarak alir (N = HV satiri sayisi). Bu, 'Distance' (0,00)
    degerlerini sertlik degerlerinden guvenilir sekilde ayirir."""
    lines = _nonempty_lines(text)
    hv_indices = [i for i, l in enumerate(lines) if HV_LABEL_PATTERN.match(l)]
    if not hv_indices:
        return []

    n = len(hv_indices)
    first_hv = hv_indices[0]

    nums = []
    j = first_hv - 1
    while j >= 0 and NUMBER_PATTERN.match(lines[j]):
        nums.insert(0, lines[j])
        j -= 1

    hardness_lines = nums[-n:] if len(nums) >= n else nums
    values = []
    for x in hardness_lines:
        try:
            values.append(float(x.replace(",", ".")))
        except ValueError:
            pass
    return values


def extract_header_fields(text):
    """MUSTERI / PARTI NO / EBAT / TARIH / OPERATOR alanlarini cikar.
    Alanlarin sirasi PDF'den PDF'e degisebildigi icin, etiket bloğu
    icindeki ve sonrasindaki deger satirlari turlerine gore siniflandirilir."""
    lines = _nonempty_lines(text)
    label_positions = [i for i, l in enumerate(lines) if l in LABELS]

    if not label_positions:
        return {"musteri_parti": None, "ebat_pdf": None, "tarih": None, "operator": None}

    first_label = min(label_positions)
    last_label = max(label_positions)

    end_idx = len(lines)
    for i in range(last_label + 1, len(lines)):
        if "SingleMeasurement" in lines[i] or re.match(r"^\d+[,\.]\d+\s*mm$", lines[i]):
            end_idx = i
            break

    content = [l for l in lines[first_label:end_idx] if l not in LABELS]

    tarih = None
    operator = None
    ebat_val = None
    leftover = []

    for l in content:
        if DATE_PATTERN.match(l):
            tarih = l
        elif l.upper() in KNOWN_OPERATORS:
            operator = l
        elif EBAT_PATTERN.match(l) and ebat_val is None:
            ebat_val = l
        else:
            leftover.append(l)

    musteri_parti = " / ".join(leftover) if leftover else None

    return {
        "musteri_parti": musteri_parti,
        "ebat_pdf": ebat_val,
        "tarih": tarih,
        "operator": operator,
    }


def parse_pdf(pdf_path):
    """Tek bir PDF dosyasindan tum alanlari cikarir."""
    text = extract_pdf_text(pdf_path)
    hv_values = extract_hv_values(text)
    fields = extract_header_fields(text)

    avg_hv = round(sum(hv_values) / len(hv_values), 2) if hv_values else None

    # Ebat -> dosya adindan turetilir (en guvenilir kaynak operatorun kendi adlandirmasi)
    ebat = Path(pdf_path).stem

    # Dosya degisiklik tarihi/saati -> isletim sistemi meta verisi
    mtime = os.path.getmtime(pdf_path)
    dt = datetime.datetime.fromtimestamp(mtime)
    tarih_str = dt.strftime("%d.%m.%Y")
    saat_str = dt.strftime("%H:%M")

    return {
        "ebat": ebat,
        "avg_hv": avg_hv,
        "musteri_parti": fields["musteri_parti"] or "",
        "operator": fields["operator"] or "",
        "dosya_adi": Path(pdf_path).name,
        "tarih": tarih_str,
        "saat": saat_str,
        "hv_count": len(hv_values),
    }


# ---------------------------------------------------------------------------
# Excel Islemleri
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
ALT_FILL = PatternFill("solid", fgColor="DCE6F1")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
THIN = Side(style="thin", color="9DC3E6")
THICK = Side(style="medium", color="1F3864")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_BORDER = Border(left=THICK, right=THICK, top=THICK, bottom=THICK)


def ensure_workbook(xlsx_path):
    """Excel dosyasi yoksa basligi ile birlikte olusturur."""
    if os.path.exists(xlsx_path):
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Sertlik Tablosu"
    col_widths = [32, 22, 28, 22, 38, 22, 20]
    for col, (h, w) in enumerate(zip(HEADERS, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = HEADER_BORDER
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    wb.save(xlsx_path)


def read_existing_filenames(xlsx_path):
    wb = load_workbook(xlsx_path)
    ws = wb.active
    names = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and len(row) >= 5 and row[4]:
            names.add(row[4])
    return names


def append_row(xlsx_path, row_data):
    wb = load_workbook(xlsx_path)
    ws = wb.active
    next_row = ws.max_row + 1

    values = [
        row_data["ebat"],
        row_data["avg_hv"],
        row_data["musteri_parti"],
        row_data["operator"],
        row_data["dosya_adi"],
        row_data["tarih"],
        row_data["saat"],
    ]

    fill = ALT_FILL if next_row % 2 == 0 else WHITE_FILL
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=next_row, column=col, value=val)
        cell.font = Font(name="Arial", size=10)
        cell.fill = fill
        cell.border = CELL_BORDER
        if col == 2:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if val is not None:
                cell.number_format = "0.00"
        elif col in (6, 7):
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[next_row].height = 18
    wb.save(xlsx_path)


# ---------------------------------------------------------------------------
# Arayuz (GUI)
# ---------------------------------------------------------------------------

class SertlikApp:
    def __init__(self, root):
        self.root = root
        self.root.title(APP_NAME)
        self.root.geometry("760x520")
        self.root.minsize(640, 420)

        self.cfg = load_config()
        self.xlsx_path = self.cfg.get("xlsx_path")

        self._build_ui()

        if not self.xlsx_path or not os.path.exists(self.xlsx_path):
            self.root.after(200, self.choose_excel_file)
        else:
            self._set_status(f"Aktif tablo: {self.xlsx_path}")

    # -- UI kurulumu -------------------------------------------------------

    def _build_ui(self):
        top = tk.Frame(self.root, bg="#1F3864", height=70)
        top.pack(fill="x", side="top")
        title = tk.Label(top, text="Sertlik Takip", font=("Arial", 18, "bold"),
                          fg="white", bg="#1F3864")
        title.pack(side="left", padx=20, pady=15)

        change_btn = tk.Button(top, text="Excel Dosyasi Degistir", command=self.choose_excel_file,
                                bg="#3B5998", fg="white", relief="flat", padx=10, pady=5,
                                font=("Arial", 10))
        change_btn.pack(side="right", padx=20, pady=15)

        self.status_label = tk.Label(self.root, text="", font=("Arial", 9), fg="#444",
                                      anchor="w", justify="left", wraplength=700)
        self.status_label.pack(fill="x", padx=20, pady=(10, 0))

        # Surukle-birak alani
        self.drop_frame = tk.Frame(self.root, bg="#EAF1FB", highlightbackground="#1F3864",
                                    highlightthickness=2, bd=0)
        self.drop_frame.pack(fill="both", expand=True, padx=20, pady=15)

        self.drop_label = tk.Label(
            self.drop_frame,
            text="PDF dosyalarini buraya surukleyip birakin\n\n(veya tiklayip secin)",
            font=("Arial", 14), bg="#EAF1FB", fg="#1F3864", justify="center"
        )
        self.drop_label.pack(expand=True)
        self.drop_label.bind("<Button-1>", lambda e: self.choose_pdf_files())
        self.drop_frame.bind("<Button-1>", lambda e: self.choose_pdf_files())

        if DND_AVAILABLE:
            self.drop_frame.drop_target_register(DND_FILES)
            self.drop_frame.dnd_bind("<<Drop>>", self._on_drop)
            self.drop_label.drop_target_register(DND_FILES)
            self.drop_label.dnd_bind("<<Drop>>", self._on_drop)

        # Log alani
        log_frame = tk.Frame(self.root)
        log_frame.pack(fill="both", expand=False, padx=20, pady=(0, 15))
        tk.Label(log_frame, text="Islem Gecmisi:", font=("Arial", 10, "bold")).pack(anchor="w")

        self.log_text = tk.Text(log_frame, height=8, font=("Consolas", 9), bg="#F7F7F7", wrap="word")
        self.log_text.pack(fill="both", expand=True)
        self.log_text.config(state="disabled")

    def _set_status(self, msg):
        self.status_label.config(text=msg)

    def _log(self, msg):
        self.log_text.config(state="normal")
        timestamp = datetime.datetime.now().strftime("%H:%M:%S")
        self.log_text.insert("end", f"[{timestamp}] {msg}\n")
        self.log_text.see("end")
        self.log_text.config(state="disabled")

    # -- Excel dosyasi secimi ----------------------------------------------

    def choose_excel_file(self):
        answer = messagebox.askyesno(
            APP_NAME,
            "Mevcut bir Excel dosyasi mi kullanmak istiyorsunuz?\n\n"
            "'Evet' -> var olan dosyayi secin\n"
            "'Hayir' -> yeni bir dosya olusturulsun"
        )
        if answer:
            path = filedialog.askopenfilename(
                title="Sertlik tablosu Excel dosyasini secin",
                filetypes=[("Excel Dosyasi", "*.xlsx")]
            )
        else:
            path = filedialog.asksaveasfilename(
                title="Yeni Excel dosyasini nereye kaydedelim?",
                defaultextension=".xlsx",
                initialfile="Sertlik_Tablosu.xlsx",
                filetypes=[("Excel Dosyasi", "*.xlsx")]
            )

        if not path:
            if not self.xlsx_path:
                messagebox.showwarning(APP_NAME, "Bir Excel dosyasi secmeden uygulamayi kullanamazsiniz.")
            return

        try:
            ensure_workbook(path)
        except Exception as e:
            messagebox.showerror(APP_NAME, f"Excel dosyasi olusturulamadi:\n{e}")
            return

        self.xlsx_path = path
        self.cfg["xlsx_path"] = path
        save_config(self.cfg)
        self._set_status(f"Aktif tablo: {path}")
        self._log(f"Excel dosyasi ayarlandi: {path}")

    # -- Dosya secimi / surukle-birak ---------------------------------------

    def choose_pdf_files(self):
        paths = filedialog.askopenfilenames(
            title="PDF dosyalarini secin",
            filetypes=[("PDF Dosyasi", "*.pdf")]
        )
        if paths:
            self.process_files(list(paths))

    def _on_drop(self, event):
        raw = event.data
        paths = self.root.tk.splitlist(raw)
        pdf_paths = [p for p in paths if p.lower().endswith(".pdf")]
        if not pdf_paths:
            messagebox.showwarning(APP_NAME, "Sadece PDF dosyalari kabul edilir.")
            return
        self.process_files(pdf_paths)

    # -- Ana islem -----------------------------------------------------------

    def process_files(self, pdf_paths):
        if not self.xlsx_path:
            messagebox.showwarning(APP_NAME, "Once bir Excel dosyasi secmelisiniz.")
            self.choose_excel_file()
            if not self.xlsx_path:
                return

        if not os.path.exists(self.xlsx_path):
            ensure_workbook(self.xlsx_path)

        existing_names = read_existing_filenames(self.xlsx_path)

        added = 0
        skipped = 0
        failed = 0

        for pdf_path in pdf_paths:
            fname = Path(pdf_path).name
            try:
                if fname in existing_names:
                    proceed = messagebox.askyesno(
                        APP_NAME,
                        f"'{fname}' dosyasi tabloda zaten kayitli.\n\nYine de eklensin mi?"
                    )
                    if not proceed:
                        self._log(f"ATLANDI (mukerrer): {fname}")
                        skipped += 1
                        continue

                row_data = parse_pdf(pdf_path)
                append_row(self.xlsx_path, row_data)
                existing_names.add(fname)
                added += 1
                self._log(
                    f"EKLENDI: {fname} | Ebat={row_data['ebat']} | "
                    f"Ort.Sertlik={row_data['avg_hv']} | Operator={row_data['operator'] or '-'} | "
                    f"{row_data['tarih']} {row_data['saat']}"
                )
            except Exception as e:
                failed += 1
                self._log(f"HATA: {fname} -> {e}")

        summary = f"Tamamlandi: {added} eklendi, {skipped} atlandi, {failed} hata."
        self._set_status(f"Aktif tablo: {self.xlsx_path}  |  {summary}")
        if added or failed:
            messagebox.showinfo(APP_NAME, summary)


def main():
    if DND_AVAILABLE:
        root = TkinterDnD.Tk()
    else:
        root = tk.Tk()
        messagebox.showwarning(
            APP_NAME,
            "Surukle-birak kutuphanesi bulunamadi.\n"
            "Dosya secmek icin pencereye tiklayin."
        )
    app = SertlikApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
